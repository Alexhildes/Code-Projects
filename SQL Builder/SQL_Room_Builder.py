
#Python Room BIM Wizard Test Program AH 2018

import openpyxl
from openpyxl import load_workbook
import argparse

headers = []

firstQuery = 'update dbo.room set '

roomData = load_workbook('E:\\Perth Childrens Hospital\\BIM\\Room Information\\ELV-RM-Combined 2018.xlsx')
f = open('E:\\Code Projects\\SQL Builder\\Output SQL\\dbo_room.sql','w')

activeSheet = roomData["BIM Link Sheet"]
maxRows = activeSheet.max_row + 1
maxColumns = activeSheet.max_column + 1
print(maxRows, 'Rooms')
print(maxColumns, 'Columns')
print('use bim',file=f)

for j in range (2, maxRows):

    headers =[]
    for i in range (1, maxColumns):
        
        
        headers.append(str(activeSheet.cell(row=j, column=i).value))
        headers = ['' if h == 'None' else h for h in headers]

        endQuery1 = '' 
        endQuery1 = "','".join(headers[1:9])
        
    print(firstQuery +
          "[floor] = '" + headers[1] +
          "', [roomdesc] = '" + headers[2] +
          "', [wayfindingname] =  '" + headers[3] +
          "', [coordref] =  '" + headers[4] +
          "', [location] = geometry::STGeomFromText('POINT (" + headers[4].replace(",","") + ")',0)" +
          ", [lightingaddress] = '" + headers[5] +
          "', [securityzone1] = '" + headers[6] +
          "', [wayfindingdept] = '" + headers[7] +
          "', [Category Number] = " + headers[8] +

          " where [roomnumber] = '" + headers[0] +
          "' if @@rowcount = 0 insert into dbo.room (" +
         
          "[roomnumber], " +
          "[floor], " +
          "[roomdesc], " +
          "[wayfindingname], " +
          "[coordref], " +
          "[lightingaddress], " +
          "[securityzone1], " +
          "[wayfindingdept], " +
          "[Category Number], " +
          "[location]) values ('" +
          headers[0] + "','" + endQuery1 +
          "'," + "geometry::STGeomFromText('POINT (" + headers[4].replace(",","") + ")',0));", file=f)
            
print("done")
#important, always close the file when you have finished

f.close()
