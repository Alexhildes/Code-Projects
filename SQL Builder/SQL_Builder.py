#Python BIM Wizard V2 Program AH 2019

#02.07.2019
#Added Checks for ID, Co-ordinate Reference, Category Number fields

import openpyxl
from openpyxl import load_workbook
import argparse
import sys

#Blank array
headers = []

#First part of the Update Query
firstQuery = 'update [dbo].[device] set '

#Loads excel file - must be in the correct format
bimData = load_workbook('.\\SQL Builder\\Source Excel Docs\\Assorted.xlsx')
#Opens blank sql file
with open('.\\SQL Builder\\Output SQL\\Nagios Updates Pre 2020.sql','w') as f:

    activeSheet = bimData.active
    maxRows = activeSheet.max_row + 1
    maxColumns = activeSheet.max_column + 1
    print(str(maxRows - 2) + ' Rows to be updated into the Database')

    #--------------------------------------------------------------#
    #Check entire column of BIM ID for Duplicates - column 1
    print('Checking ID....')
    bimID = []
    for i in range (2, maxRows):
        if activeSheet.cell(row=i, column=1).value in bimID and activeSheet.cell(row=i, column=1).value != None:
            print('ID is not unique at Row ' + str(i) + ', ensure each Record has a unique ID. Consult Points Lists or dbo.device SQL table and try again')
            #Exit the script
            sys.exit()
        else:
            bimID.append(activeSheet.cell(row=i, column=1).value)
    
    print('ID correct')
    #--------------------------------------------------------------#
    #Check entire column of Co-ordinate reference - column 7
    print('Checking Cordinate Reference....')
    coordindateRef = []
    for i in range (2, maxRows):
        if activeSheet.cell(row=i, column=7).value in coordindateRef:
            pass # if already in list do nothing
        else:
            coordindateRef.append(str(activeSheet.cell(row=i, column=7).value))
            coordindateRef = ['' if h == 'None' else h for h in coordindateRef] #Any blank cells will be replaced with ''

    if all(', ' in x for x in coordindateRef):
        print('Co-ordinate Reference correct')
    else:
        print('Co-ordinate Reference incorrect, format must be "xxx.xxx, yyy.yyy" only, correct the import spreadsheet and try again')
        #Exit the script
        sys.exit()
    #--------------------------------------------------------------#
    #Checks Category Number is between 1 and 5
    print('Checking Category Number....')
    categorynumber = []
    #Check entire column of Category Number - column 18 and create a array of unique values
    for i in range (2, maxRows):
        if activeSheet.cell(row=i, column=18).value in categorynumber:
            pass # if already in list do nothing
        else:
            categorynumber.append(activeSheet.cell(row=i, column=18).value)
            categorynumber = [0 if h == None else h for h in categorynumber] #Any blank cells will be replaced with 0
    #Check if category number is between 1 and 5
    if (all(1 <= x <= 5 for x in categorynumber)):
        print('Category Number is correct')
    else:
        print('Catgory Number must be between 1 and 5, correct the import spreadsheet and try again')
        #Exit the script
        sys.exit()
    #--------------------------------------------------------------#

    print('Creating SQL File')
    print('use [bim]',file=f)
    print('',file=f)
    print('GO', file=f)
    print('',file=f)

    for j in range (2, maxRows):

        if j % 100 == 0:
            #Breaks the query up into transactions of 100
            print('',file=f)
            print('GO',file=f)
            print('',file=f)

        headers =[]
        for i in range (1, maxColumns):

            headers.append(str(activeSheet.cell(row=j, column=i).value))
            headers = ['' if h == 'None' else h for h in headers]

            endQuery1 = '' 
            endQuery1 = "','".join(headers[1:32])

        print(firstQuery +
              "[modelname] = '" + headers[1] +
              "', [Name] = '" + headers[2] +
              "', [Mark] =  '" + headers[3] +
              "', [Department] = '" + headers[4] +
              "', [wayfindingdept] = '" + headers[5] +
              "', [Coordinate Reference] = '" + headers[6] +
              "', [Room Number] = '" + headers[7] +
              "', [Room Name] = '" + headers[8] +
              "', [wayfindingname] = '" + headers[9] +
              "', [Floor] = '" + headers[10] +
              "', [CCTV Primary Target] = '" + headers[11] +
              "', [CCTV Ancillary Target 1] = '" + headers[12] +
              "', [CCTV Ancillary Target 2] = '" + headers[13] +
              "', [Associated CCTV Camera] = '" + headers[14] +
              "', [Functional Area] = '" + headers[15] +
              "', [Category] = '" + headers[16] +
              "', [Category Number] = " + headers[17] +
              ", [Security Zone 1] = '" + headers[18] +
              "', [Security Zone 2] = '" + headers[19] +
              "', [HVAC Zone] = '" + headers[20] +
              "', [Fire Alarm Zone] = '" + headers[21] +
              "', [Fire Alarm Zone 2] = '" + headers[22] +
              "', [System Owner] = '" + headers[23] +
              "', [Termination Destination] = '" + headers[24] +
              "', [Point Name] = '" + headers[25] +
              "', [Parent Name] = '" + headers[26] +
              "', [IP Address] = '" + headers[27] +
              "', [Lighting Control Address 1] = '" + headers[28] +
              "', [Lighting Control Address 2] = '" + headers[29] +
              "', [Display by default] = '" + headers[30] +
              "', [Family and Type] = '" + headers[31] +
              "', [type] = '" + headers[33] +
              "', [assetid] = '" + headers[34] +
              "', [assetname] = '" + headers[35] +
              "', [assetnumber] = '" + headers[36] +
              "', [location] = geometry::STGeomFromText('POINT (" + headers[6].replace(",","") + ")',0)" +

              " where [Id] = " + str(headers[0]) +
              " if @@rowcount = 0 insert into device (" +

              "[Id], " +
              "[modelname], " +
              "[Name], " +
              "[Mark], " +
              "[Department], " +
              "[wayfindingdept], " +
              "[Coordinate Reference], " +
              "[Room Number], " +
              "[Room Name], " +
              "[wayfindingname], " +
              "[Floor], " +
              "[CCTV Primary Target], " +
              "[CCTV Ancillary Target 1], " +
              "[CCTV Ancillary Target 2], " +
              "[Associated CCTV Camera], " +
              "[Functional Area], " +
              "[Category], " +
              "[Category Number], " +
              "[Security Zone 1], " +
              "[Security Zone 2], " +
              "[HVAC Zone], " +
              "[Fire Alarm Zone], " +
              "[Fire Alarm Zone 2], " +
              "[System Owner], " +
              "[Termination Destination], " +
              "[Point Name], " +
              "[Parent Name], " +
              "[IP Address], " +
              "[Lighting Control Address 1], " +
              "[Lighting Control Address 2], " +
              "[Display by Default], " +
              "[Family and Type], " +
              "[type], " +
              "[assetid], " +
              "[assetname], " +
              "[assetnumber], "
              "[location]) values (" +
              headers[0] + ",'" + endQuery1 +
              "','" + headers[33] +
              "','" + headers[34] +
              "','" + headers[35] +
              "','" + headers[36] +
              "'," + "geometry::STGeomFromText('POINT (" + headers[6].replace(",","") + ")',0));", file=f)

    print('',file=f)
    print("GO",file=f)

print("SQL Insert File Created")
