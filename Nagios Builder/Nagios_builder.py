import openpyxl
from openpyxl import load_workbook

import argparse
stencil = load_workbook('.\\\\Nagios Builder\\\\Book1.xlsx')
with open('.\\\\Nagios Builder\\\\Ward 5A Cameras.cfg','w') as f:    
    print(stencil.sheetnames)
    ws = stencil['Nagios']
    maxRows = ws.max_row + 1
    
    print(maxRows)
    
    for i in range(1, maxRows):
        print('define host{' + '\\n'            
        'use PCH-appliance-host-template ' + '\\n'
        'host_name ' + str(ws.cell(row=i, column=1).value) + '\\n'
        'hostgroups cctv_cameras_hostgroup, CCTV, Camera, SourceSystemID_SC.008,IME219-1VI ' + '\\n'
        'display_name ' + str(ws.cell(row=i, column=1).value) + ' - IP Camera ' + str(ws.cell(row=i, column=1).value)[3:6] + '\\n'
        'alias CCTV - ' + str(ws.cell(row=i, column=1).value) + ' - IP Camera ' + str(ws.cell(row=i, column=1).value)[3:6] + '\\n'
        'notes location= ' + '\\n'
        'address ' + str(ws.cell(row=i, column=2).value) + '\\n'
        '}' + '\\n',file=f)
